"""导出模块

包含Excel表格和图片导出功能。
"""

import io
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from typing import Dict, List, Tuple, Optional
import zipfile

# 设置matplotlib中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


def export_assignment_to_excel(
    assignment: Dict[str, int], 
    seats: List[Tuple[int, int]], 
    n_cols: int, 
    rows_per_col: List[int]
) -> bytes:
    """将座位分配方案导出为Excel
    
    Args:
        assignment: 座位分配方案
        seats: 座位列表
        n_cols: 列数
        rows_per_col: 每列的排数列表
        
    Returns:
        bytes: Excel文件的字节数据
    """
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "座位表"
    
    # 写入标题行
    ws.merge_cells(f'A1:{chr(65+n_cols)}1')
    ws.cell(1, 1).value = "座位安排表"
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    
    # 写入列标题
    for c in range(n_cols):
        ws.cell(2, c + 2).value = f"列 {c+1}"
        ws.cell(2, c + 2).alignment = Alignment(horizontal='center')
    
    # 写入行标题
    max_rows = max(rows_per_col) if rows_per_col else 0
    for r in range(max_rows):
        ws.cell(r + 3, 1).value = f"排 {r+1}"
        ws.cell(r + 3, 1).alignment = Alignment(horizontal='center')
    
    # 创建座位ID到人名的映射
    seat_to_person = {}
    for name, idx in assignment.items():
        seat_to_person[idx] = name
    
    # 填充座位信息
    for idx, (c, r) in enumerate(seats):
        name = seat_to_person.get(idx, "")
        ws.cell(r + 3, c + 2).value = name
        ws.cell(r + 3, c + 2).alignment = Alignment(horizontal='center')
        
        # 添加单元格颜色
        if name:
            fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            ws.cell(r + 3, c + 2).fill = fill
    
    # 调整列宽
    for c in range(1, n_cols + 2):
        ws.column_dimensions[chr(64 + c)].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_assignment_to_image(
    assignment: Dict[str, int], 
    seats: List[Tuple[int, int]], 
    n_cols: int, 
    rows_per_col: List[int], 
    pair_weights: Optional[Dict[frozenset, float]] = None,
    figsize: Optional[Tuple[float, float]] = None,
    dpi: int = 120,
    split_visualization: bool = False,
    aisles: Optional[List[int]] = None,
    show_all_lines: bool = False
) -> bytes:
    """绘制座位图，带关系线
    
    Args:
        assignment: 座位分配方案
        seats: 座位列表
        n_cols: 列数
        rows_per_col: 每列的排数列表
        pair_weights: 人员对权重字典（可选，用于绘制关系线）
        figsize: 图片尺寸（可选）
        dpi: 图片分辨率
        split_visualization: 是否分离可视化
        aisles: 过道位置列表（可选，列索引）
        
    Returns:
        bytes: PNG图片的字节数据
    """
    max_rows = max(rows_per_col) if rows_per_col else 1
    cell_w, cell_h = 2.0, 1.2
    margin = 0.5
    
    if figsize is None:
        fig_w = max(8, n_cols * cell_w + 2 * margin)
        fig_h = max(6, max_rows * cell_h + 2 * margin)
        figsize = (fig_w, fig_h)
    
    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    fig.patch.set_facecolor('#f8f9fa')
    ax.set_xlim(0, n_cols * cell_w + 2 * margin)
    ax.set_ylim(0, max_rows * cell_h + 2 * margin)
    ax.axis('off')
    
    # 创建座位ID到人名的映射
    seat_to_person = {}
    person_to_pos = {}
    
    # 画从左到右，底部为R1，向上增加
    for idx, (c, r) in enumerate(seats):
        x0 = margin + c * cell_w
        y0 = margin + (max_rows - r - 1) * cell_h  # 反转Y轴，使R1在底部
        
        # 座位矩形和编号
        rect_color = '#e9ecef'
        rect = patches.Rectangle((x0, y0), cell_w, cell_h, fill=True, 
                               linewidth=1, edgecolor='#adb5bd', facecolor=rect_color)
        ax.add_patch(rect)
        seat_id = f"C{c+1}-R{r+1}"
        ax.text(x0 + 0.1, y0 + cell_h - 0.2, seat_id, ha='left', va='top', 
                fontsize=7, color='#6c757d')
        
        # 找到这个座位的人
        name = None
        for k, v in assignment.items():
            if v == idx:
                name = k
                seat_to_person[idx] = name
                person_to_pos[name] = (x0 + cell_w/2, y0 + cell_h/2)
                break
        
        # 显示人名
        label = name if name else ""
        if label:
            ax.text(x0 + cell_w/2, y0 + cell_h/2, label, ha='center', va='center', 
                    fontsize=min(9, max(6, 60//len(label))), fontweight='bold')
    
    # 绘制过道线条
    if aisles:
        for aisle_col in aisles:
            if 0 <= aisle_col < n_cols:
                # 在两列之间绘制过道线，位置在列间隙的中心
                x_aisle = margin + (aisle_col + 1) * cell_w
                ax.axvline(x=x_aisle, ymin=margin/(max_rows * cell_h + 2 * margin), 
                          ymax=(margin + max_rows * cell_h)/(max_rows * cell_h + 2 * margin),
                          color='#6c757d', linewidth=3, linestyle='--', alpha=0.7)
                # 添加过道标识
                ax.text(x_aisle + 0.1, margin + max_rows * cell_h / 2, '过道', 
                       rotation=90, ha='left', va='center', fontsize=8, color='#6c757d')
    
    # 如果提供了权重信息，画出关系线
    if pair_weights:
        # 创建座位位置映射
        seat_positions = {idx: (c, r) for idx, (c, r) in enumerate(seats)}
        
        def are_seats_adjacent(person1, person2):
            """检查两个人的座位是否相邻"""
            # 找到两个人的座位索引
            seat1_idx = assignment.get(person1)
            seat2_idx = assignment.get(person2)
            
            if seat1_idx is None or seat2_idx is None:
                return False
                
            # 获取座位坐标
            if seat1_idx not in seat_positions or seat2_idx not in seat_positions:
                return False
                
            c1, r1 = seat_positions[seat1_idx]
            c2, r2 = seat_positions[seat2_idx]
            
            # 检查是否相邻（包括对角线）
            dc = abs(c1 - c2)
            dr = abs(r1 - r2)
            
            return (dc <= 1 and dr <= 1) and (dc + dr > 0)
        
        # 收集所有连线信息
        if show_all_lines:
            # 显示所有关系线
            pos_pairs = [(list(p)[0], list(p)[1], w) for p, w in pair_weights.items() 
                        if w > 3.0 and list(p)[0] in person_to_pos and list(p)[1] in person_to_pos]
            neg_pairs = [(list(p)[0], list(p)[1], w) for p, w in pair_weights.items() 
                        if w < -3.0 and list(p)[0] in person_to_pos and list(p)[1] in person_to_pos]
        else:
            # 只包含实际满足的关系（相邻座位）
            pos_pairs = [(list(p)[0], list(p)[1], w) for p, w in pair_weights.items() 
                        if w > 3.0 and list(p)[0] in person_to_pos and list(p)[1] in person_to_pos 
                        and are_seats_adjacent(list(p)[0], list(p)[1])]
            neg_pairs = [(list(p)[0], list(p)[1], w) for p, w in pair_weights.items() 
                        if w < -3.0 and list(p)[0] in person_to_pos and list(p)[1] in person_to_pos 
                        and not are_seats_adjacent(list(p)[0], list(p)[1])]
        
        # 如果启用拆分可视化，返回两张图的字节数据
        if split_visualization:
            return _create_split_visualization(assignment, seats, n_cols, rows_per_col, 
                                             pos_pairs, neg_pairs, person_to_pos, 
                                             figsize, dpi, margin, cell_w, cell_h, max_rows, aisles, show_all_lines)
        
        # 统计每个位置的连线数量，用于颜色区分
        from collections import defaultdict
        pos_count = defaultdict(int)
        neg_count = defaultdict(int)
        
        # 预统计连线数量
        for a, b, w in pos_pairs:
            pos_count[a] += 1
            pos_count[b] += 1
        for a, b, w in neg_pairs:
            neg_count[a] += 1
            neg_count[b] += 1
        
        # 定义颜色列表
        pos_colors = ['#28a745', '#20c997', '#17a2b8', '#6f42c1', '#e83e8c']
        neg_colors = ['#dc3545', '#fd7e14', '#ffc107', '#6c757d', '#343a40']
        
        # 绘制正向关系线
        pos_drawn = defaultdict(int)
        for a, b, w in pos_pairs:
            if a in person_to_pos and b in person_to_pos:
                x1, y1 = person_to_pos[a]
                x2, y2 = person_to_pos[b]
                
                # 根据已绘制的线数选择颜色和偏移
                color_idx = (pos_drawn[a] + pos_drawn[b]) % len(pos_colors)
                color = pos_colors[color_idx]
                
                # 计算偏移量，避免线条重叠
                offset = (pos_drawn[a] + pos_drawn[b]) * 0.05
                dx = (y2 - y1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                dy = -(x2 - x1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                
                # 绘制连线
                ax.plot([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                       color=color, alpha=0.6, linewidth=2, solid_capstyle='round')
                
                # 在连线两端添加圆点
                ax.scatter([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                          c=color, s=30, alpha=0.8, zorder=5)
                
                pos_drawn[a] += 1
                pos_drawn[b] += 1
        
        # 绘制负向关系线
        neg_drawn = defaultdict(int)
        for a, b, w in neg_pairs:
            if a in person_to_pos and b in person_to_pos:
                x1, y1 = person_to_pos[a]
                x2, y2 = person_to_pos[b]
                
                # 根据已绘制的线数选择颜色和偏移
                color_idx = (neg_drawn[a] + neg_drawn[b]) % len(neg_colors)
                color = neg_colors[color_idx]
                
                # 计算偏移量，避免线条重叠
                offset = (neg_drawn[a] + neg_drawn[b]) * 0.05
                dx = (y2 - y1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                dy = -(x2 - x1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                
                # 绘制虚线
                ax.plot([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                       color=color, alpha=0.6, linewidth=1.5, linestyle='--', solid_capstyle='round')
                
                # 在连线两端添加方形标记
                ax.scatter([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                          c=color, s=25, alpha=0.8, marker='s', zorder=5)
                
                neg_drawn[a] += 1
                neg_drawn[b] += 1
    
    # 添加标题和图例
    ax.text(margin, max_rows * cell_h + margin/2, "座位分布图", 
            fontsize=12, fontweight='bold')
    
    # 添加改进的图例
    if pair_weights:
        # 创建图例元素
        from matplotlib.lines import Line2D
        legend_elements = []
        
        if pos_pairs:
            legend_elements.append(Line2D([0], [0], color='#28a745', linewidth=2, 
                                        marker='o', markersize=6, label='喜欢坐一起（实线+圆点）'))
        if neg_pairs:
            legend_elements.append(Line2D([0], [0], color='#dc3545', linewidth=1.5, 
                                        linestyle='--', marker='s', markersize=5, label='不喜欢坐一起（虚线+方点）'))
        
        if legend_elements:
            ax.legend(handles=legend_elements, loc='upper right', fontsize=8, 
                     framealpha=0.9, fancybox=True, shadow=True)
            
        # 添加颜色说明
        if len(pos_pairs) > 1 or len(neg_pairs) > 1:
            ax.text(0.02, 0.98, '注：多条线重叠时使用不同颜色区分', 
                   transform=ax.transAxes, fontsize=7, verticalalignment='top',
                   bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8))
    
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _create_split_visualization(
    assignment: Dict[str, int], 
    seats: List[Tuple[int, int]], 
    n_cols: int, 
    rows_per_col: List[int],
    pos_pairs: List[Tuple[str, str, float]],
    neg_pairs: List[Tuple[str, str, float]],
    person_to_pos: Dict[str, Tuple[float, float]],
    figsize: Tuple[float, float],
    dpi: int,
    margin: float,
    cell_w: float,
    cell_h: float,
    max_rows: int,
    aisles: Optional[List[int]] = None,
    show_all_lines: bool = False
) -> bytes:
    """创建分离的正负关系可视化图
    
    Returns:
        bytes: ZIP文件的字节数据，包含两张PNG图片
    """
    from collections import defaultdict
    
    # 创建座位位置映射
    seat_positions = {idx: (c, r) for idx, (c, r) in enumerate(seats)}
    
    def are_seats_adjacent(person1, person2):
        """检查两个人的座位是否相邻"""
        # 找到两个人的座位索引
        seat1_idx = assignment.get(person1)
        seat2_idx = assignment.get(person2)
        
        if seat1_idx is None or seat2_idx is None:
            return False
            
        # 获取座位坐标
        if seat1_idx not in seat_positions or seat2_idx not in seat_positions:
            return False
            
        c1, r1 = seat_positions[seat1_idx]
        c2, r2 = seat_positions[seat2_idx]
        
        # 检查是否相邻（包括对角线）
        dc = abs(c1 - c2)
        dr = abs(r1 - r2)
        
        return (dc <= 1 and dr <= 1) and (dc + dr > 0)
    
    # 过滤关系对，只保留满足条件的
    filtered_pos_pairs = [(a, b, w) for a, b, w in pos_pairs if are_seats_adjacent(a, b)]
    filtered_neg_pairs = [(a, b, w) for a, b, w in neg_pairs if not are_seats_adjacent(a, b)]
    
    # 创建ZIP文件
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 只创建正向关系图
        if filtered_pos_pairs:
            fig1, ax1 = plt.subplots(figsize=figsize, dpi=dpi)
            fig1.patch.set_facecolor('#f8f9fa')
            ax1.set_xlim(0, n_cols * cell_w + 2 * margin)
            ax1.set_ylim(0, max_rows * cell_h + 2 * margin)
            ax1.axis('off')
            
            # 绘制座位
            _draw_seats(ax1, assignment, seats, margin, cell_w, cell_h, max_rows, aisles)
            
            # 绘制正向关系线
            pos_colors = ['#28a745', '#20c997', '#17a2b8', '#6f42c1', '#e83e8c']
            pos_drawn = defaultdict(int)
            
            for a, b, w in filtered_pos_pairs:
                if a in person_to_pos and b in person_to_pos:
                    x1, y1 = person_to_pos[a]
                    x2, y2 = person_to_pos[b]
                    
                    color_idx = (pos_drawn[a] + pos_drawn[b]) % len(pos_colors)
                    color = pos_colors[color_idx]
                    
                    offset = (pos_drawn[a] + pos_drawn[b]) * 0.05
                    dx = (y2 - y1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                    dy = -(x2 - x1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                    
                    ax1.plot([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                           color=color, alpha=0.6, linewidth=2, solid_capstyle='round')
                    ax1.scatter([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                              c=color, s=30, alpha=0.8, zorder=5)
                    
                    pos_drawn[a] += 1
                    pos_drawn[b] += 1
            
            ax1.text(margin, max_rows * cell_h + margin/2, "座位分布图 - 正向关系", 
                    fontsize=12, fontweight='bold')
            
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            
            buf1 = io.BytesIO()
            plt.savefig(buf1, format="png", dpi=dpi, bbox_inches="tight", facecolor=fig1.get_facecolor())
            plt.close(fig1)
            buf1.seek(0)
            zip_file.writestr("positive_relationships.png", buf1.getvalue())
        
        # 创建负向关系图
        if filtered_neg_pairs:
            fig2, ax2 = plt.subplots(figsize=figsize, dpi=dpi)
            fig2.patch.set_facecolor('#f8f9fa')
            ax2.set_xlim(0, n_cols * cell_w + 2 * margin)
            ax2.set_ylim(0, max_rows * cell_h + 2 * margin)
            ax2.axis('off')
            
            # 绘制座位
            _draw_seats(ax2, assignment, seats, margin, cell_w, cell_h, max_rows, aisles)
            
            # 绘制负向关系线
            neg_colors = ['#dc3545', '#fd7e14', '#ffc107', '#6c757d', '#343a40']
            neg_drawn = defaultdict(int)
            
            for a, b, w in filtered_neg_pairs:
                if a in person_to_pos and b in person_to_pos:
                    x1, y1 = person_to_pos[a]
                    x2, y2 = person_to_pos[b]
                    
                    color_idx = (neg_drawn[a] + neg_drawn[b]) % len(neg_colors)
                    color = neg_colors[color_idx]
                    
                    offset = (neg_drawn[a] + neg_drawn[b]) * 0.05
                    dx = (y2 - y1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                    dy = -(x2 - x1) * offset / max(1, abs(x2 - x1) + abs(y2 - y1))
                    
                    ax2.plot([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                           color=color, alpha=0.6, linewidth=1.5, linestyle='--', solid_capstyle='round')
                    ax2.scatter([x1 + dx, x2 + dx], [y1 + dy, y2 + dy], 
                              c=color, s=25, alpha=0.8, marker='s', zorder=5)
                    
                    neg_drawn[a] += 1
                    neg_drawn[b] += 1
            
            ax2.text(margin, max_rows * cell_h + margin/2, "座位分布图 - 负向关系", 
                    fontsize=12, fontweight='bold')
            
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            
            buf2 = io.BytesIO()
            plt.savefig(buf2, format="png", dpi=dpi, bbox_inches="tight", facecolor=fig2.get_facecolor())
            plt.close(fig2)
            buf2.seek(0)
            zip_file.writestr("negative_relationships.png", buf2.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def _draw_seats(
    ax, 
    assignment: Dict[str, int], 
    seats: List[Tuple[int, int]], 
    margin: float, 
    cell_w: float, 
    cell_h: float, 
    max_rows: int,
    aisles: Optional[List[int]] = None
):
    """绘制座位的辅助函数"""
    for idx, (c, r) in enumerate(seats):
        x0 = margin + c * cell_w
        y0 = margin + (max_rows - r - 1) * cell_h
        
        rect_color = '#e9ecef'
        rect = patches.Rectangle((x0, y0), cell_w, cell_h, fill=True, 
                               linewidth=1, edgecolor='#adb5bd', facecolor=rect_color)
        ax.add_patch(rect)
        
        seat_id = f"C{c+1}-R{r+1}"
        ax.text(x0 + 0.1, y0 + cell_h - 0.2, seat_id, ha='left', va='top', 
                fontsize=7, color='#6c757d')
        
        # 找到这个座位的人
        name = None
        for k, v in assignment.items():
            if v == idx:
                name = k
                break
        
        # 显示人名
        label = name if name else ""
        if label:
            ax.text(x0 + cell_w/2, y0 + cell_h/2, label, ha='center', va='center', 
                    fontsize=min(9, max(6, 60//len(label))), fontweight='bold')
    
    # 绘制过道线条
    if aisles:
        n_cols = max(c for c, r in seats) + 1 if seats else 0
        for aisle_col in aisles:
            if 0 <= aisle_col < n_cols - 1:  # 确保过道在有效范围内
                # 在指定列的右侧绘制过道线，位置在两列之间的中心
                x_aisle = margin + (aisle_col + 1) * cell_w
                y_start = margin
                y_end = margin + max_rows * cell_h
                
                ax.axvline(x=x_aisle, ymin=y_start/(max_rows * cell_h + 2 * margin),
                          ymax=y_end/(max_rows * cell_h + 2 * margin),
                          color='#6c757d', linewidth=3, linestyle='--', alpha=0.7)
                # 添加过道标识
                ax.text(x_aisle + cell_w * 0.1, margin + max_rows * cell_h / 2, '过道', 
                       rotation=90, ha='left', va='center', fontsize=8, color='#6c757d')


def create_assignment_summary_excel(
    results: List[Dict], 
    people: List[str], 
    seats: List[Tuple[int, int]]
) -> bytes:
    """创建包含多个方案对比的Excel文件
    
    Args:
        results: 座位分配结果列表
        people: 人员列表
        seats: 座位列表
        
    Returns:
        bytes: Excel文件的字节数据
    """
    wb = openpyxl.Workbook()
    
    # 创建摘要工作表
    summary_ws = wb.active
    summary_ws.title = "方案对比"
    
    # 写入摘要表头
    headers = ["方案编号", "目标函数值", "满足率(%)", "满足人数", "满足对数", "求解状态"]
    for col, header in enumerate(headers, 1):
        summary_ws.cell(1, col).value = header
        summary_ws.cell(1, col).alignment = Alignment(horizontal='center')
    
    # 写入摘要数据
    for i, result in enumerate(results, 1):
        summary_ws.cell(i + 1, 1).value = f"方案{i}"
        summary_ws.cell(i + 1, 2).value = round(result.get('objective', 0), 2)
        summary_ws.cell(i + 1, 3).value = result.get('satisfaction_rate', 0)
        summary_ws.cell(i + 1, 4).value = result.get('n_satisfied', 0)
        summary_ws.cell(i + 1, 5).value = result.get('n_satisfied_pairs', 0)
        summary_ws.cell(i + 1, 6).value = "最优解" if result.get('status') == 4 else "可行解"
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        summary_ws.column_dimensions[chr(64 + col)].width = 15
    
    # 为每个方案创建详细工作表
    for i, result in enumerate(results, 1):
        ws = wb.create_sheet(f"方案{i}详情")
        assignment = result.get('assignment', {})
        
        # 写入详细分配信息
        ws.cell(1, 1).value = "姓名"
        ws.cell(1, 2).value = "座位号"
        ws.cell(1, 1).alignment = Alignment(horizontal='center')
        ws.cell(1, 2).alignment = Alignment(horizontal='center')
        
        for row, (name, seat_idx) in enumerate(assignment.items(), 2):
            ws.cell(row, 1).value = name
            if 0 <= seat_idx < len(seats):
                c, r = seats[seat_idx]
                ws.cell(row, 2).value = f"C{c+1}-R{r+1}"
            else:
                ws.cell(row, 2).value = "未知座位"
            
            ws.cell(row, 1).alignment = Alignment(horizontal='center')
            ws.cell(row, 2).alignment = Alignment(horizontal='center')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_layout_preview(
    n_cols: int, 
    rows_per_col: List[int], 
    edges: List[Tuple[Tuple[int, int], Tuple[int, int]]],
    figsize: Optional[Tuple[float, float]] = None,
    dpi: int = 100
) -> bytes:
    """导出教室布局预览图
    
    Args:
        n_cols: 列数
        rows_per_col: 每列的排数列表
        edges: 邻座边列表
        figsize: 图片尺寸（可选）
        dpi: 图片分辨率
        
    Returns:
        bytes: PNG图片的字节数据
    """
    max_rows = max(rows_per_col) if rows_per_col else 0
    
    if figsize is None:
        figsize = (max(8, n_cols * 1.2), max(6, max_rows * 1.2))
    
    fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
    
    # 绘制座位
    for c in range(n_cols):
        for r in range(rows_per_col[c]):
            rect = patches.Rectangle((c, r), 0.9, 0.9, fill=True, 
                                   edgecolor='black', facecolor='lightgray', alpha=0.5)
            ax.add_patch(rect)
            ax.text(c + 0.45, r + 0.45, f"C{c+1}-R{r+1}", ha='center', va='center', fontsize=8)
    
    # 绘制邻座关系
    for (c1, r1), (c2, r2) in edges:
        ax.plot([c1 + 0.45, c2 + 0.45], [r1 + 0.45, r2 + 0.45], 'b-', alpha=0.3)
    
    ax.set_xlim(-0.5, n_cols + 0.5)
    ax.set_ylim(-0.5, max_rows + 0.5)
    ax.set_xticks(range(n_cols))
    ax.set_yticks(range(max_rows))
    ax.set_xticklabels([f'C{i+1}' for i in range(n_cols)])
    ax.set_yticklabels([f'R{i+1}' for i in range(max_rows)])
    ax.set_title('教室座位布局预览')
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()