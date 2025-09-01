"""数据处理模块

包含Excel文件读取、数据解析和预处理功能。
"""

import pandas as pd
import openpyxl
from typing import List, Dict, Tuple, Set, Optional
from collections import defaultdict
import streamlit as st


def load_names_from_excel(file, sheet_name: Optional[str] = None, name_col_spec: str = "A:A") -> List[str]:
    """从Excel文件加载人员名单
    
    Args:
        file: 上传的文件对象
        sheet_name: 工作表名称，None表示使用第一个工作表
        name_col_spec: 名单列范围，如"A:A"
        
    Returns:
        List[str]: 人员名单列表
    """
    try:
        # 确保文件对象有效
        if file is None:
            st.error("文件对象为空")
            return []
            
        # 读取Excel文件
        try:
            # 如果没有指定sheet_name，使用0来明确指定第一个工作表
            # 这样可以避免pd.read_excel返回字典的情况
            if sheet_name is None:
                df = pd.read_excel(file, sheet_name=0, header=None, usecols=name_col_spec)
            else:
                df = pd.read_excel(file, sheet_name=sheet_name, header=None, usecols=name_col_spec)
        except Exception as read_error:
            st.error(f"读取Excel文件失败: {str(read_error)}")
            return []
            
        # 检查DataFrame类型
        if not isinstance(df, pd.DataFrame):
            # 如果返回字典，尝试转换为DataFrame
            if isinstance(df, dict):
                try:
                    # 尝试从字典创建DataFrame
                    df = pd.DataFrame(df)
                    st.warning("Excel文件格式异常，已尝试自动修复")
                except Exception as convert_error:
                    st.error(f"无法转换数据格式: {str(convert_error)}")
                    return []
            else:
                st.error(f"读取的数据类型错误: {type(df)}，期望DataFrame。请检查Excel文件格式。")
                return []
            
        # 检查DataFrame是否有效
        if df is None or df.empty:
            st.error("Excel文件为空或无法读取")
            return []
            
        # 获取第一列的所有非空值，跳过第一行（表头）
        names = df.iloc[1:, 0].dropna().astype(str).tolist()
        # 去重并保持顺序
        names = list(dict.fromkeys(names))
        # 过滤空字符串和表头关键词
        names = [name.strip() for name in names if name.strip() and name.strip().lower() not in ['姓名', 'name', '名字', '学生姓名']]
        return names
    except Exception as e:
        import traceback
        error_msg = f"加载名单出错: {str(e)}\n详细错误:\n{traceback.format_exc()}"
        st.error(error_msg)
        return []


def parse_cell_range(range_spec: str) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """解析单元格范围，如 'A1:C5'
    
    Args:
        range_spec: 单元格范围字符串
        
    Returns:
        Tuple: (start_col_idx, start_row_idx, end_col_idx, end_row_idx)
    """
    if not range_spec or ":" not in range_spec:
        return None, None, None, None
    
    try:
        start, end = range_spec.split(":")
        
        # 解析起始单元格
        start_col = ""
        start_row = ""
        for c in start:
            if c.isalpha():
                start_col += c
            else:
                start_row += c
        
        # 解析结束单元格
        end_col = ""
        end_row = ""
        for c in end:
            if c.isalpha():
                end_col += c
            else:
                end_row += c
        
        # 转换列标签为数字索引（0-based）
        start_col_idx = openpyxl.utils.column_index_from_string(start_col) - 1
        end_col_idx = openpyxl.utils.column_index_from_string(end_col) - 1
        
        # 转换行标签为数字索引（0-based）
        start_row_idx = int(start_row) - 1 if start_row else 0
        end_row_idx = int(end_row) - 1 if end_row else None
        
        return start_col_idx, start_row_idx, end_col_idx, end_row_idx
    except Exception:
        return None, None, None, None


def auto_detect_preference_ranges(ws) -> Tuple[Optional[str], Optional[str]]:
    """自动识别喜好名单的单元格范围
    
    Args:
        ws: openpyxl工作表对象
        
    Returns:
        Tuple[Optional[str], Optional[str]]: (willing_range, unwilling_range)
    """
    try:
        # 扫描整个工作表找到有数据的区域（第一行开始就是有效数据）
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row == 1 and max_col == 1:
            return None, None
            
        # 找到所有有数据的列
        data_columns = []
        for col in range(1, max_col + 1):
            has_data = False
            for row in range(1, max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                data_columns.append(col)
        
        if len(data_columns) < 2:
            return None, None
            
        # 识别空列分隔，找到愿意和不愿意区域
        willing_cols = []
        unwilling_cols = []
        
        # 查找连续的数据列组
        groups = []
        current_group = [data_columns[0]]
        
        for i in range(1, len(data_columns)):
            if data_columns[i] - data_columns[i-1] == 1:
                # 连续列
                current_group.append(data_columns[i])
            else:
                # 有间隔，开始新组
                groups.append(current_group)
                current_group = [data_columns[i]]
        groups.append(current_group)
        
        # 假设第一组是愿意，第二组是不愿意
        if len(groups) >= 1:
            willing_cols = groups[0]
        if len(groups) >= 2:
            unwilling_cols = groups[1]
            
        # 为每组找到最长的列来确定行数（从第一行开始计算）
        def find_max_row_in_cols(cols):
            max_data_row = 1
            for col in cols:
                for row in range(max_row, 0, -1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        max_data_row = max(max_data_row, row)
                        break
            return max_data_row
            
        willing_range = None
        unwilling_range = None
        
        if willing_cols:
            willing_max_row = find_max_row_in_cols(willing_cols)
            willing_start_col = min(willing_cols)
            willing_end_col = max(willing_cols)
            willing_range = f"{chr(64 + willing_start_col)}1:{chr(64 + willing_end_col)}{willing_max_row}"
            
        if unwilling_cols:
            unwilling_max_row = find_max_row_in_cols(unwilling_cols)
            unwilling_start_col = min(unwilling_cols)
            unwilling_end_col = max(unwilling_cols)
            unwilling_range = f"{chr(64 + unwilling_start_col)}1:{chr(64 + unwilling_end_col)}{unwilling_max_row}"
            
        return willing_range, unwilling_range
        
    except Exception as e:
        return None, None


def load_preferences_from_excel(
    file, 
    sheet_name: Optional[str] = None, 
    willing_range_spec: str = "A1:B10", 
    unwilling_range_spec: str = "D1:E10",
    auto_detect: bool = False
) -> Tuple[Dict[int, Set], Dict[int, Set], List[Tuple], List[Tuple], Optional[Tuple]]:
    """从Excel加载喜好/不喜好关系
    
    Args:
        file: 上传的文件对象
        sheet_name: 工作表名称
        willing_range_spec: 喜好关系单元格范围
        unwilling_range_spec: 不喜好关系单元格范围
        auto_detect: 是否自动识别单元格范围
        
    Returns:
        Tuple: (willing_pairs_by_rank, unwilling_pairs_by_rank, willing_headers, unwilling_headers)
    """
    willing_pairs_by_rank = defaultdict(set)
    unwilling_pairs_by_rank = defaultdict(set)
    
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 如果启用自动检测，则自动识别范围
        detection_results = None
        if auto_detect:
            detected_willing, detected_unwilling = auto_detect_preference_ranges(ws)
            if detected_willing:
                willing_range_spec = detected_willing
            if detected_unwilling:
                unwilling_range_spec = detected_unwilling
            detection_results = (detected_willing, detected_unwilling)
        
        willing_headers = []
        unwilling_headers = []
        
        # 解析喜好关系单元格范围
        w_start_col, w_start_row, w_end_col, w_end_row = parse_cell_range(willing_range_spec)
        if w_start_col is not None:
            rank = 1
            # 动态识别实际的等级数量：每一列为一个等级
            total_cols = w_end_col - w_start_col + 1
            max_levels = total_cols
            
            for level in range(max_levels):
                col = w_start_col + level
                if col <= w_end_col:
                    # 读取该列的所有人名对（逗号分隔格式），从第一行开始
                    pairs_found = False
                    for row in range(w_start_row + 1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col + 1).value
                        if cell_value and str(cell_value).strip():
                            # 解析逗号分隔的人名对
                            parts = str(cell_value).strip().split(',')
                            if len(parts) == 2:
                                a, b = parts[0].strip(), parts[1].strip()
                                if a and b and a != b:
                                    willing_pairs_by_rank[rank].add(frozenset([a, b]))
                                    pairs_found = True
                    
                    # 生成标题：喜欢等级X:列名
                    col_letter = openpyxl.utils.get_column_letter(col + 1)
                    header = f"喜欢等级{rank}:{col_letter}列"
                    willing_headers.append((header, ""))
                    # 只有找到实际数据对时才增加等级计数
                    if pairs_found:
                        rank += 1
                    else:
                        # 即使没有数据也要增加等级计数以保持一致性
                        rank += 1
        
        # 解析不喜好关系单元格范围
        uw_start_col, uw_start_row, uw_end_col, uw_end_row = parse_cell_range(unwilling_range_spec)
        if uw_start_col is not None:
            rank = 1
            # 动态识别实际的等级数量：每一列为一个等级
            total_cols = uw_end_col - uw_start_col + 1
            max_levels = total_cols
            
            for level in range(max_levels):
                col = uw_start_col + level
                if col <= uw_end_col:
                    # 读取该列的所有人名对（逗号分隔格式），从第一行开始
                    pairs_found = False
                    for row in range(uw_start_row + 1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col + 1).value
                        if cell_value and str(cell_value).strip():
                            # 解析逗号分隔的人名对
                            parts = str(cell_value).strip().split(',')
                            if len(parts) == 2:
                                a, b = parts[0].strip(), parts[1].strip()
                                if a and b and a != b:
                                    unwilling_pairs_by_rank[rank].add(frozenset([a, b]))
                                    pairs_found = True
                    
                    # 生成标题：不喜欢等级X:列名
                    col_letter = openpyxl.utils.get_column_letter(col + 1)
                    header = f"不喜欢等级{rank}:{col_letter}列"
                    unwilling_headers.append((header, ""))
                    # 只有找到实际数据对时才增加等级计数
                    if pairs_found:
                        rank += 1
                    else:
                        # 即使没有数据也要增加等级计数以保持一致性
                        rank += 1
        
        wb.close()
    except Exception as e:
        st.error(f"读取Excel出错: {str(e)}")
        return {}, {}, [], [], None
    
    return willing_pairs_by_rank, unwilling_pairs_by_rank, willing_headers, unwilling_headers, detection_results


def parse_custom_weights(text: str) -> List[Tuple[str, str, float]]:
    """解析自定义权重文本
    
    Args:
        text: 自定义权重文本，每行格式为"人名1,人名2,权重"
        
    Returns:
        List[Tuple[str, str, float]]: 自定义权重对列表
    """
    if not text:
        return []
    
    custom_pairs = []
    lines = text.strip().split('\n')
    for line in lines:
        parts = line.split(',')
        if len(parts) == 3:
            try:
                name1 = parts[0].strip()
                name2 = parts[1].strip()
                weight = float(parts[2].strip())
                if name1 and name2:
                    custom_pairs.append((name1, name2, weight))
            except ValueError:
                continue
    return custom_pairs


def compute_pair_weights(
    willing_pairs_by_rank: Dict[int, Set], 
    unwilling_pairs_by_rank: Dict[int, Set],
    w_weights: List[float], 
    uw_weights: List[float], 
    custom_pairs: Optional[List[Tuple[str, str, float]]] = None
) -> Dict[frozenset, float]:
    """计算所有对的权重
    
    Args:
        willing_pairs_by_rank: 按等级分组的喜好关系
        unwilling_pairs_by_rank: 按等级分组的不喜好关系
        w_weights: 喜好权重列表
        uw_weights: 不喜好权重列表
        custom_pairs: 自定义权重对
        
    Returns:
        Dict[frozenset, float]: 权重字典
    """
    pair_weights = {}
    
    # 处理喜好对
    for rank, pairs in willing_pairs_by_rank.items():
        if rank <= len(w_weights):
            for pair in pairs:
                pair_weights[pair] = w_weights[rank-1]
    
    # 处理不喜好对
    for rank, pairs in unwilling_pairs_by_rank.items():
        if rank < len(uw_weights):
            for pair in pairs:
                # 不喜欢关系使用负权重（uw_weights已经是负数）
                pair_weights[pair] = uw_weights[rank]
    
    # 添加自定义权重对
    if custom_pairs:
        for name1, name2, weight in custom_pairs:
            pair = frozenset([name1, name2])
            pair_weights[pair] = float(weight)
    
    return pair_weights